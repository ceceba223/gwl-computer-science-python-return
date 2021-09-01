import os.path
import time

from forex_python.converter import CurrencyRates 
from forex_python.bitcoin import BtcConverter

from openpyxl import Workbook, load_workbook
from datetime import date, datetime

def save_to_xlsx(BTC_USD, BTC_EUR, USD_EUR):
    workbook = load_workbook('data.xlsx')
    sheet = workbook.active

    date = str(datetime.now())

    row = (date, BTC_USD, BTC_EUR, USD_EUR)
    sheet.append(row)

    workbook.save(filename="data.xlsx")

def setup_xlsx():
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "date"
    sheet["B1"] = "1 BTC in USD"
    sheet["C1"] = "1 BTC in EUR"
    sheet["D1"] = "1 USD in EUR"

    workbook.save(filename="data.xlsx")

def usd_in_eur():
    c = CurrencyRates()
    EUR = c.get_rate('USD', 'EUR')  #convert USD to EURO
    return(EUR)

def btc_in_usd():
    b = BtcConverter()
    USD = b.get_latest_price('USD')  #convert BTC to USD
    return(USD)

def btc_in_eur():
    b = BtcConverter()
    EUR = b.get_latest_price('EUR')   #convert BTC to EUR
    return(EUR)


def main():
    if os.path.isfile('data.xlsx'):
        pass
    else:
        setup_xlsx()

    while True:
        BTC_USD = round(btc_in_usd(), 2)
        USD_EUR = round(usd_in_eur(), 2)
        BTC_EUR = round(btc_in_eur(), 2)

        save_to_xlsx(BTC_USD, BTC_EUR, USD_EUR)

        # Sleep for 24 hours 
        # (For testing purposes you can set it to a lower number)
        time.sleep(24 * 60 * 60)

if __name__ == '__main__':
    main()